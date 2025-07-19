from datetime import date
from enum import Enum
from pathlib import Path
import requests
import time

from bs4 import BeautifulSoup, Tag, Comment
import pandas as pd
from openpyxl import load_workbook, Workbook

from .caching import CACHE_DIR, load_file_cache, cache_file, handle_request
from .html_parsing import fetch_sub_tag, fetch_data_stat, fetch_specific_sub_tag
from .user_prompting import prompt_for_stat

pd.set_option("future.no_silent_downcasting", True)


class Positions(Enum):
    """Position enum for more self-documenting and type-safe position refs"""

    QB = "QB"
    WR = "WR"
    RB = "RB"
    TE = "TE"


class Team:
    """Fetches information about play percentage and rate-of-play over the last 3 years for a given team"""

    def __init__(
        self, team_name: str, use_cache: bool = True, save_results: bool = True
    ):
        self.team_name = team_name
        self.current_year = date.today().year
        index = pd.Index(range(self.current_year - 3, self.current_year + 1))
        columns = [
            "head_coach",
            "offensive_coordinator",
            "total_plays",
            "run_percent",
            "run_plays",
            "pass_percent",
            "pass_plays",
        ]
        dtypes = {
            "head_coach": "string",
            "offensive_coordinator": "string",
            "total_plays": "int32",
            "run_percent": "float32",
            "run_plays": "int32",
            "pass_percent": "float32",
            "pass_plays": "int32",
        }
        self.historical_data = pd.DataFrame(index=index, columns=columns)
        for year in index:
            team_path = CACHE_DIR / self.team_name / f"team_{year}.html"
            if use_cache and team_path.exists():
                team_data = load_file_cache(
                    team_path, f"Using cache for {team_name} stats for {year}..."
                )
                doc = BeautifulSoup(team_data, "html.parser")
            else:
                response = handle_request(
                    f"https://pro-football-reference.com/teams/{team_name}/{year}.htm",
                    request_message=f"Looking up {team_name} stats for {year}...",
                )
                if save_results:
                    cache_file(
                        team_path,
                        response.text,
                        f"Saving {team_name} stats for {year}...",
                    )
                doc = BeautifulSoup(response.text, "html.parser")
            if year != self.current_year:
                table_body = fetch_sub_tag(doc, "tbody")
                table_row = fetch_sub_tag(table_body, "tr")

                run_plays = fetch_data_stat(
                    table_row, "Rushing attempts", "rush_att", stat_dtype=int
                )
                assert type(run_plays) == int

                pass_plays = fetch_data_stat(
                    table_row, "Passing attempts", "pass_att", stat_dtype=int
                )
                assert type(pass_plays) == int

                total_plays: int = run_plays + pass_plays
                self.historical_data.loc[year, "total_plays"] = total_plays
                self.historical_data.loc[year, "run_percent"] = round(
                    run_plays / total_plays * 100, 2
                )
                self.historical_data.loc[year, "pass_percent"] = round(
                    pass_plays / total_plays * 100, 2
                )
                self.historical_data.loc[year, "run_plays"] = run_plays
                self.historical_data.loc[year, "pass_plays"] = pass_plays
            for p in doc.find_all("p"):
                assert type(p) == Tag
                if "Coach:" in p.get_text():
                    head_coach_html = fetch_sub_tag(p, "a")
                    self.historical_data.loc[year, "head_coach"] = (
                        head_coach_html.get_text()
                    )
                elif "Offensive Coordinator:" in p.get_text():
                    oc_html = fetch_sub_tag(p, "a")
                    self.historical_data.loc[year, "offensive_coordinator"] = (
                        oc_html.get_text()
                    )

        self.historical_data = self.historical_data.fillna(0).astype(dtypes)

    def project(self):
        """Prompts the user to enter projections for the current season"""
        print(self.historical_data)
        print(
            "Average plays: ",
            self.historical_data["total_plays"].drop(self.current_year).mean(),
        )
        print(
            "Average rushing play percentage: ",
            self.historical_data["run_percent"].drop(self.current_year).mean(),
        )
        print(
            "Average passing play percentage: ",
            self.historical_data["pass_percent"].drop(self.current_year).mean(),
        )
        print()

        self.total_plays = prompt_for_stat("total plays", int, self.current_year)

        self.run_percent = 0
        self.pass_percent = 0
        while self.run_percent + self.pass_percent != 100:
            print(
                "Run plays + pass plays do not add up to 100% of total plays. Please re-enter run/pass percent data"
            )
            self.run_percent = prompt_for_stat(
                "rushing play percentage", int, self.current_year
            )
            assert type(self.run_percent) == int

            self.pass_percent = prompt_for_stat(
                "passing play percentage", int, self.current_year
            )
            assert type(self.pass_percent) == int

    def save_projections(self, filename: str):
        """Saves your team-level projection to a sheet"""
        assert type(self.total_plays) == int
        assert type(self.run_percent) == int
        assert type(self.pass_percent) == int
        file_path = Path(filename)
        file_path.parent.mkdir(parents=True, exist_ok=True)
        if not file_path.exists():
            workbook = Workbook()
            workbook.create_sheet(self.team_name.capitalize())
            workbook.save(file_path)
            current_index = 1
        else:
            wb = load_workbook(filename)
            if self.team_name.capitalize() not in wb.sheetnames:
                wb.create_sheet(self.team_name.capitalize())
                wb.save(filename)
            existing_data = pd.read_excel(
                filename, sheet_name=self.team_name.capitalize()
            )
            current_index = existing_data.shape[0] + 1
        formatted_data = pd.DataFrame()
        formatted_data.loc[current_index, "Total Plays"] = self.total_plays
        formatted_data.loc[current_index, "Run Plays"] = round(
            self.total_plays * self.run_percent / 100, 2
        )
        formatted_data.loc[current_index, "Pass Plays"] = round(
            self.total_plays * self.pass_percent / 100, 2
        )
        if current_index > 1:
            print(existing_data)
            df_combined = pd.concat([existing_data, formatted_data])
            print(df_combined)
            with pd.ExcelWriter(
                filename, engine="openpyxl", mode="a", if_sheet_exists="replace"
            ) as writer:
                print("Merging data in excel...")
                df_combined.to_excel(
                    writer, sheet_name=self.team_name.capitalize(), index=False
                )
        else:
            with pd.ExcelWriter(
                filename, engine="openpyxl", mode="a", if_sheet_exists="replace"
            ) as writer:
                print("Saving new data to excel...")
                formatted_data.to_excel(
                    writer, sheet_name=self.team_name.capitalize(), index=False
                )

    def get_player(
        self,
        requested_position: Positions,
        use_cache: bool = True,
        save_results: bool = True,
    ) -> tuple[str | None, str | None]:
        """Return a player to project given an HTML roster and position type to parse"""
        # Check if team data already cached
        roster_path = CACHE_DIR / self.team_name / f"roster_{self.current_year}.html"
        # Pull from cache if using and cache hit
        if use_cache and roster_path.exists():
            roster_data = load_file_cache(
                roster_path,
                f"Using cache for {self.current_year} team roster for {self.team_name}...",
            )
            doc = BeautifulSoup(roster_data, "html.parser")
        else:
            # Fetch roster data from ProFootballReference if not cached
            response = handle_request(
                f"https://pro-football-reference.com/teams/{self.team_name}/{self.current_year}_roster.htm",
                f"Fetching {self.current_year} team roster for {self.team_name}...",
            )

            if save_results:
                cache_file(
                    roster_path,
                    response.text,
                    f"Saving {self.current_year} team roster for {self.team_name}...",
                )

            doc = BeautifulSoup(response.text, "html.parser")
        players = self._find_players(doc, requested_position)
        return self._select_player(players)

    def _find_players(
        self, doc: Tag, requested_position: Positions
    ) -> list[tuple[str, Tag]]:
        player_div = fetch_specific_sub_tag(doc, "div", "all_roster")

        # Find table of players
        table_doc = None
        for d in player_div.descendants:
            if isinstance(d, Comment):
                table_doc = BeautifulSoup(d.string, "html.parser")

        if table_doc is None:
            raise ValueError("Table not found")

        table_body = fetch_sub_tag(table_doc, "tbody")
        table_rows = table_body.find_all("tr")
        print(type(table_rows))
        if table_rows is None:
            print(table_body)
            raise ValueError("Table rows not found")

        print(f"Finding all {requested_position.value}s for {self.team_name}")
        player_name = None
        players = []
        for row in table_rows:
            if type(row) is not Tag:
                print(row)
                print(type(row))
                raise TypeError("Table row does not have expected type. Check HTML")

            position_string = fetch_data_stat(row, "position", "pos")
            if position_string not in Positions:
                # Log that this position isn't supported
                continue

            position = Positions(position_string)
            if position != requested_position:
                continue

            player_data = fetch_data_stat(
                row, "player name", "player", return_html=True
            )
            assert type(player_data) == Tag

            player_name = player_data.get_text()
            print(player_name)
            players.append((player_name, player_data))

        return players

    def _select_player(self, players: list[tuple[str, Tag]]):
        for player_name, player_data in players:
            decision = (
                input(f"Would you like to do projections for {player_name}? ").lower()
                == "y"
            )
            if decision:
                player_attribute = fetch_sub_tag(player_data, "a")
                player_url_suffix = player_attribute.get("href")
                player_url = f"https://pro-football-reference.com{player_url_suffix}"
                return player_name, player_url

        return None, None
