from datetime import date
from pathlib import Path
import requests
import time
from bs4 import BeautifulSoup, Comment, Tag
import pandas as pd
from openpyxl import load_workbook
from .teams import Team

class QB:
    """Projection process for quarterbacks"""
    def __init__(self, team: Team):
        """Searches through QBs on a team's current roster until you want to project one, then collects historical data for the past 3 years about that player"""
        self.team = team
        self.current_year = date.today().year
        index = pd.Index(range(self.current_year - 3, self.current_year + 1))
        columns = ['team',
                   'games played',
                   'games started',
                   'pass_att',
                   'int %',
                   'pass td %',
                   'comp %',
                   'rush %',
                   'ypc',
                   'tds/rush_yard'
                   ]
        self.historical_data = pd.DataFrame(index=index, columns=columns)
        team_url = f'https://pro-football-reference.com/teams/{self.team.team_name}/{self.current_year}_roster.htm'
        print(f'Fetching {self.current_year} team roster for {self.team.team_name}...')
        response = requests.get(team_url)
        time.sleep(6)
        if response.status_code != 200:
            raise requests.HTTPError(f'Request unsuccessful. Response code: {response.status_code}')
        doc = BeautifulSoup(response.text, 'html.parser')
        
        player_div: Tag = doc.find('div', id='all_roster')
        
        for d in player_div.descendants:
            if isinstance(d, Comment):
                table_doc = BeautifulSoup(d.string, 'html.parser')

        table_rows = table_doc.find('tbody').find_all('tr')

        print(f'Finding all QBs for {team.team_name}')
        for row in table_rows:
            pos = row.find(attrs={'data-stat':'pos'}).get_text()
            if pos != 'QB':
                continue
            player_data = row.find(attrs={'data-stat':'player'})
            player_name = player_data.get_text()
            print(player_name)

        for row in table_rows:
            pos = row.find(attrs={'data-stat':'pos'}).get_text()
            if pos != 'QB':
                continue
            player_data = row.find(attrs={'data-stat':'player'})
            player_name = player_data.get_text()
            decision = input(f'Would you like to do projections for {player_name}? ').lower() == 'y'
            if decision:
                self.player_name = player_name
                player_url_suffix = player_data.find('a').get('href')
                player_url = f'https://pro-football-reference.com{player_url_suffix}'
                break

        if 'player_url' not in locals():
            print('No projections will be done for this instance of QB')
            self.projections_exist = False
            return

        if player_url is None:
            raise ValueError('Player URL undefined')

        print(f'Fetching stats for QB {player_name}...')
        response = requests.get(player_url)
        time.sleep(6)
        if response.status_code != 200:
            raise requests.HTTPError(f'Request unsuccessful. Response code: {response.status_code}')
        doc = BeautifulSoup(response.text, 'html.parser')
    
        rushing_table = doc.find('table', id='rushing_and_receiving')
        if rushing_table is None:
            print('Rushing/receiving table does not exist... assuming player is rookie')
            self.projections_exist = True
            return

        for year in range(self.current_year - 3, self.current_year):
            passing_row = doc.find('table', id='passing').find('tbody').find('tr', id=f'passing.{year}')
            rushing_row = rushing_table.find('tbody').find('tr', id=f'rushing_and_receiving.{year}')
            if rushing_row is None or passing_row is None:
                print(f'{self.player_name} does not have data for {year}')
                continue
            current_stat = rushing_row.find(attrs={'data-stat':'team'}).get_text()
            if current_stat == '':
                self.historical_data.loc[year, 'team'] = 'NA'
            else:
                self.historical_data.loc[year, 'team'] = current_stat

            current_stat = rushing_row.find(attrs={'data-stat':'g'}).get_text()
            if current_stat == '':
                self.historical_data.loc[year, 'games played'] = 'NA'
            else:
                self.historical_data.loc[year, 'games played'] = int(current_stat)

            current_stat = rushing_row.find(attrs={'data-stat':'gs'}).get_text()
            if current_stat == '':
                self.historical_data.loc[year, 'games started'] = 'NA'
            else:
                self.historical_data.loc[year, 'games started'] = int(current_stat)

            self.historical_data.loc[year, 'pass_att'] = float(passing_row.find(attrs={'data-stat':'pass_att'}).get_text())
            self.historical_data.loc[year, 'int %'] = float(passing_row.find(attrs={'data-stat':'pass_int_perc'}).get_text())
            self.historical_data.loc[year, 'pass td %'] = float(passing_row.find(attrs={'data-stat':'pass_td_perc'}).get_text())
            self.historical_data.loc[year, 'comp %'] = float(passing_row.find(attrs={'data-stat':'pass_cmp_perc'}).get_text())
            self.historical_data.loc[year, 'rush %'] = float(rushing_row.find(attrs={'data-stat':'rush_att'}).get_text()) / self.team.historical_data.loc[year, 'run_plays'] * 100
            self.historical_data.loc[year, 'ypc'] = float(rushing_row.find(attrs={'data-stat':'rush_yds_per_att'}).get_text())
            self.historical_data.loc[year, 'tds/rush_yard'] = int(rushing_row.find(attrs={'data-stat':'rush_td'}).get_text()) / float(rushing_row.find(attrs={'data-stat':'rush_yds'}).get_text())

        self.projections_exist = True

    def project(self):
        """Prompts the user to enter projections for the current season"""
        print(self.historical_data)
        print('Average int %: ', self.historical_data['int %'].drop(self.current_year).mean())
        print('Average rush %: ', self.historical_data['rush %'].drop(self.current_year).mean())
        print('Average ypc: ', self.historical_data['ypc'].drop(self.current_year).mean())
        print('Average tds/rush_yard: ', self.historical_data['tds/rush_yard'].drop(self.current_year).mean())
        print()
        need_answer = True
        while need_answer:
            try:
                games_started = input('Estimated games played for 2024: ')
                games_started = int(games_started)
                need_answer = False
            except ValueError:
                print('This is not a valid integer!')
        need_answer = True
        while need_answer:
            try:
                int_percent = input('Estimated int % for 2024: ')
                int_percent = float(int_percent)
                need_answer = False
            except ValueError:
                print('This is not a valid decimal!')
        need_answer = True
        while need_answer:
            try:
                rush_percent = input('Estimated rush % for 2024: ')
                rush_percent = float(rush_percent)
                need_answer = False
            except ValueError:
                print('This is not a valid decimal!')
        need_answer = True
        while need_answer:
            try:
                ypc = input('Estimated ypc for 2024: ')
                ypc = float(ypc)
                need_answer = False
            except ValueError:
                print('This is not a valid decimal!')
        need_answer = True
        while need_answer:
            try:
                td_yard_ratio = input('Estimated tds/rush_yard for 2024: ')
                td_yard_ratio = float(td_yard_ratio)
                need_answer = False
            except ValueError:
                print('This is not a valid decimal!')

        self.games_started = games_started
        self.int_percent = int_percent
        self.rush_percent = rush_percent
        self.ypc = ypc
        self.td_yard_ratio = td_yard_ratio

    def save_projections(self, filename: str):
        """Saves your team-level projection to a sheet"""
        file_path = Path(filename)
        if file_path.parent.exists: 
            wb = load_workbook(filename)
            if self.team.team_name.capitalize() not in wb.sheetnames:
                wb.create_sheet(self.team.team_name.capitalize())
                wb.save(filename)
            existing_data = pd.read_excel(filename, sheet_name=self.team.team_name.capitalize())
            print(existing_data)
            if 'Player Name' in existing_data:
                current_index = existing_data['Player Name'].size + 1
            else:
                current_index = 2
        else:
            file_path.parent.mkdir(parents=True)
            current_index = 2
            
        formatted_data = pd.DataFrame()
        formatted_data.loc[current_index, 'Pos'] = 'QB'
        formatted_data.loc[current_index, 'Player Name'] = self.player_name
        formatted_data.loc[current_index, 'Games Started'] = self.games_started
        formatted_data.loc[current_index, 'Interception %'] = self.int_percent
        formatted_data.loc[current_index, 'Rush Share'] = self.rush_percent
        formatted_data.loc[current_index, 'Yards/Carry'] = self.ypc
        formatted_data.loc[current_index, 'TDs/Yard'] = self.td_yard_ratio
        print(formatted_data)
        if 'existing_data' in locals():
            df_combined = pd.concat([existing_data, formatted_data])
            print(df_combined)
            df_combined.to_excel(filename, sheet_name=self.team.team_name.capitalize(), index=False)
        else:
            formatted_data.to_excel(filename, sheet_name=self.team.team_name.capitalize(), index=False)

class SkillPlayer:
    """Projection process for quarterbacks"""
    def __init__(self, team: Team, pos: str):
        """Searches through QBs on a team's current roster until you want to project one, then collects historical data for the past 3 years about that player"""
        if pos not in ['RB', 'WR', 'TE']:
            raise ValueError('Position must be one of RB, WR, or TE')
        self.team = team
        self.position = pos
        self.current_year = date.today().year
        index = pd.Index(range(self.current_year - 3, self.current_year))
        columns = ['games played',
                   'games started',
                   'team',
                   'target share',
                   'catch %',
                   'yards/catch',
                   'TDs/rec_yard',
                   'rush share',
                   'ypc',
                   'tds/rush_yard'
                   ]
        self.historical_data = pd.DataFrame(index=index, columns=columns)
        team_url = f'https://pro-football-reference.com/teams/{self.team.team_name}/{self.current_year}_roster.htm'
        print(f'Fetching {self.current_year} team roster for {self.team.team_name}...')
        response = requests.get(team_url)
        time.sleep(6)
        if response.status_code != 200:
            raise requests.HTTPError(f'Request unsuccessful. Response code: {response.status_code}')
        doc = BeautifulSoup(response.text, 'html.parser')
        
        player_div: Tag = doc.find('div', id='all_roster')
        
        for d in player_div.descendants:
            if isinstance(d, Comment):
                table_doc = BeautifulSoup(d.string, 'html.parser')

        table_rows = table_doc.find('tbody').find_all('tr')

        print(f'Finding all {self.position}s for {team.team_name}')
        for row in table_rows:
            pos = row.find(attrs={'data-stat':'pos'}).get_text()
            if pos != self.position:
                continue
            player_data = row.find(attrs={'data-stat':'player'})
            player_name = player_data.get_text()
            print(player_name)

        for row in table_rows:
            pos = row.find(attrs={'data-stat':'pos'}).get_text()
            if pos != self.position:
                continue
            player_data = row.find(attrs={'data-stat':'player'})
            player_name = player_data.get_text()
            decision = input(f'Would you like to do projections for {player_name}? ').lower() == 'y'
            if decision:
                self.player_name = player_name
                player_url_suffix = player_data.find('a').get('href')
                player_url = f'https://pro-football-reference.com{player_url_suffix}'
                break

        if 'player_url' not in locals():
            print('Skipping projections for the current player instance')
            self.projections_exist = False
            return

        if player_url is None:
            raise ValueError('Player URL undefined')

        print(f'Fetching stats for {self.position} {player_name}...')
        response = requests.get(player_url)
        time.sleep(6)
        if response.status_code != 200:
            raise requests.HTTPError(f'Request unsuccessful. Response code: {response.status_code}')
        doc = BeautifulSoup(response.text, 'html.parser')
    
        rushing_table = doc.find('table', id='rushing_and_receiving')
        if rushing_table is None:
            print('Table not found under rushing_and_receiving, searching for receiving_and_rushing')
            rushing_table = doc.find('table', id='receiving_and_rushing')
            search_str = 'receiving_and_rushing'
        else:
            search_str = 'rushing_and_receiving'
        if rushing_table is None:
            print('Rushing/receiving table does not exist... assuming player is rookie')
            self.projections_exist = True
            return

        for year in range(self.current_year - 3, self.current_year):
            print(f'Finding stats for {self.player_name} in {year}...')
            rushing_row = rushing_table.find('tbody').find('tr', id=f'{search_str}.{year}')
            if rushing_row is None:
                print(f'{self.player_name} does not have data for {year}')
                continue

            current_stat = rushing_row.find(attrs={'data-stat':'team'}).get_text()
            if current_stat == '':
                self.historical_data.loc[year, 'team'] = 'NA'
            else:
                self.historical_data.loc[year, 'team'] = current_stat

            current_stat = rushing_row.find(attrs={'data-stat':'g'}).get_text()
            if current_stat == '':
                self.historical_data.loc[year, 'games played'] = 'NA'
            else:
                self.historical_data.loc[year, 'games played'] = int(current_stat)

            current_stat = rushing_row.find(attrs={'data-stat':'gs'}).get_text()
            if current_stat == '':
                self.historical_data.loc[year, 'games started'] = 'NA'
            else:
                self.historical_data.loc[year, 'games started'] = int(current_stat)

            current_stat = rushing_row.find(attrs={'data-stat':'targets'}).get_text()
            if current_stat == '':
                self.historical_data.loc[year, 'target share'] = 0
            else:
                self.historical_data.loc[year, 'target share'] = float(current_stat) / self.team.historical_data.loc[year, 'pass_plays'] * 100
            
            current_stat = rushing_row.find(attrs={'data-stat':'catch_pct'}).get_text().rstrip('%')
            if current_stat == '':
                self.historical_data.loc[year, 'catch %'] = 0
            else:
                self.historical_data.loc[year, 'catch %'] = float(current_stat)
            
            current_stat = rushing_row.find(attrs={'data-stat':'rec_yds_per_rec'}).get_text()
            if current_stat == '':
                self.historical_data.loc[year, 'yards/catch'] = 0
            else:
                self.historical_data.loc[year, 'yards/catch'] = float(current_stat)

            current_stat = rushing_row.find(attrs={'data-stat':'rec_td'}).get_text()
            if current_stat == '':
                rec_td = 0
            else:
                rec_td = int(current_stat)

            current_stat = rushing_row.find(attrs={'data-stat':'rec_yds'}).get_text()
            if current_stat == '':
                self.historical_data.loc[year, 'TDs/rec_yard'] = 0
            else:
                self.historical_data.loc[year, 'TDs/rec_yard'] = rec_td / float(current_stat)   

            current_stat = rushing_row.find(attrs={'data-stat':'rush_att'}).get_text()
            if current_stat == '':
                self.historical_data.loc[year, 'rush share'] = 0
            else:
                self.historical_data.loc[year, 'rush share'] = float(current_stat) / self.team.historical_data.loc[year, 'run_plays'] * 100

            current_stat = rushing_row.find(attrs={'data-stat':'rush_yds_per_att'}).get_text()
            if current_stat == '':
                self.historical_data.loc[year, 'ypc'] = 0
            else:
                self.historical_data.loc[year, 'ypc'] = float(current_stat)

            current_stat = rushing_row.find(attrs={'data-stat':'rush_td'}).get_text()
            if current_stat == '':
                rush_tds = 0
            else:
                rush_tds = int(current_stat)

            current_stat = rushing_row.find(attrs={'data-stat':'rush_yds'}).get_text()
            if current_stat == '':
                self.historical_data.loc[year, 'tds/rush_yard'] = 0
            else:
                self.historical_data.loc[year, 'tds/rush_yard'] = rush_tds / float(current_stat)

        self.projections_exist = True

    def project(self):
        """Prompts the user to enter projections for the current season"""
        print(self.historical_data)
        print('Average target share: ', self.historical_data['target share'].dropna(how="all").mean())
        print('Average catch %: ', self.historical_data['catch %'].dropna(how="all").mean())
        print('Average yards/catch: ', self.historical_data['yards/catch'].dropna(how="all").mean())
        print('Average TDs/rec_yard: ', self.historical_data['TDs/rec_yard'].dropna(how="all").mean())
        print('Average rush share: ', self.historical_data['rush share'].dropna(how="all").mean())
        print('Average ypc: ', self.historical_data['ypc'].dropna(how="all").mean())
        print('Average tds/rush_yard: ', self.historical_data['tds/rush_yard'].dropna(how="all").mean())
        print()
        need_answer = True
        while need_answer:
            try:
                games_started = input('Estimated games played for 2024: ')
                games_started = int(games_started)
                need_answer = False
            except ValueError:
                print('This is not a valid integer!')
        need_answer = True
        while need_answer:
            try:
                target_share = input('Estimated target share for 2024: ')
                target_share = float(target_share)
                need_answer = False
            except ValueError:
                print('This is not a valid decimal!')
        need_answer = True
        while need_answer:
            try:
                catch_percent = input('Estimated catch % for 2024: ')
                catch_percent = float(catch_percent)
                need_answer = False
            except ValueError:
                print('This is not a valid decimal!')
        need_answer = True
        while need_answer:
            try:
                yds_per_catch = input('Estimated yards per catch for 2024: ')
                yds_per_catch = float(yds_per_catch)
                need_answer = False
            except ValueError:
                print('This is not a valid decimal!')
        need_answer = True
        while need_answer:
            try:
                tds_per_rec_yd = input('Estimated TDs/rec_yard for 2024: ')
                tds_per_rec_yd = float(tds_per_rec_yd)
                need_answer = False
            except ValueError:
                print('This is not a valid decimal!')
        need_answer = True
        while need_answer:
            try:
                rush_percent = input('Estimated rush share for 2024: ')
                rush_percent = float(rush_percent)
                need_answer = False
            except ValueError:
                print('This is not a valid decimal!')
        need_answer = True
        while need_answer:
            try:
                ypc = input('Estimated ypc for 2024: ')
                ypc = float(ypc)
                need_answer = False
            except ValueError:
                print('This is not a valid decimal!')
        need_answer = True
        while need_answer:
            try:
                td_yard_ratio = input('Estimated tds/rush_yard for 2024: ')
                td_yard_ratio = float(td_yard_ratio)
                need_answer = False
            except ValueError:
                print('This is not a valid decimal!')

        self.games_started = games_started
        self.target_share = target_share
        self.catch_percent = catch_percent
        self.yds_per_catch = yds_per_catch
        self.tds_per_rec_yd = tds_per_rec_yd
        self.rush_percent = rush_percent
        self.ypc = ypc
        self.td_yard_ratio = td_yard_ratio

    def save_projections(self, filename: str):
        """Saves your player projection to a sheet"""
        file_path = Path(filename)
        if file_path.parent.exists: 
            wb = load_workbook(filename)
            if self.team.team_name.capitalize() not in wb.sheetnames:
                wb.create_sheet(self.team.team_name.capitalize())
                wb.save(filename)
            existing_data = pd.read_excel(filename, sheet_name=self.team.team_name.capitalize())
            if 'Player Name' in existing_data:
                current_index = existing_data['Player Name'].size + 1
            else:
                current_index = 2
        else:
            file_path.parent.mkdir(parents=True)
            current_index = 2
            
        formatted_data = pd.DataFrame()
        formatted_data.loc[current_index, 'Pos'] = self.position
        formatted_data.loc[current_index, 'Player Name'] = self.player_name
        formatted_data.loc[current_index, 'Games Started'] = self.games_started
        formatted_data.loc[current_index, 'Target Share'] = self.target_share
        formatted_data.loc[current_index, 'Catch Percentage'] = self.catch_percent
        formatted_data.loc[current_index, 'Yards/Catch'] = self.yds_per_catch
        formatted_data.loc[current_index, 'TDs/receiving yard'] = self.tds_per_rec_yd
        formatted_data.loc[current_index, 'Rush Share'] = self.rush_percent
        formatted_data.loc[current_index, 'Yards/Carry'] = self.ypc
        formatted_data.loc[current_index, 'TDs/Yard'] = self.td_yard_ratio
        print(formatted_data)
        if 'existing_data' in locals():
            df_combined = pd.concat([existing_data, formatted_data])
            print(df_combined)
            df_combined.to_excel(filename, sheet_name=self.team.team_name.capitalize(), index=False)
        else:
            formatted_data.to_excel(filename, sheet_name=self.team.team_name.capitalize(), index=False)
