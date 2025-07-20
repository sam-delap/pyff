"""
Implementation of projection utils for skill-position players (non-QBs)
"""

from datetime import date
from pathlib import Path
from bs4 import BeautifulSoup, Tag
import pandas as pd
from openpyxl import load_workbook
from .teams import Team, Positions
from .caching import CACHE_DIR, load_file_cache, cache_file, handle_request
from .html_parsing import fetch_sub_tag, fetch_data_stat
from .user_prompting import prompt_for_stat


class SkillPlayer:
    """Parent class for non-quarterbacks"""

    def __init__(
        self,
        team: Team,
        position: Positions,
        use_cache: bool = True,
        save_results: bool = True,
    ):
        """Searches through skill players on a team's current roster until you want to project one, then collects historical data for the past 3 years about that player"""
        if position not in [Positions.RB, Positions.TE, Positions.WR]:
            raise ValueError("Position must be one of RB, WR, or TE")
        self.team = team
        self.position = position
        self.current_year = date.today().year
        index = pd.Index(range(self.current_year - 3, self.current_year))
        columns = [
            "games played",
            "games started",
            "team",
            "target share",
            "catch %",
            "yards/catch",
            "TDs/rec_yard",
            "rush share",
            "ypc",
            "tds/rush_yard",
        ]
        self.historical_data = pd.DataFrame(index=index, columns=columns)
        player_name, player_url = team.get_player(position)

        if player_name is None or player_url is None:
            print(f"No projections will be done for this instance of {position.value}")
            self.projections_exist = False
            return

        self.player_name = player_name
        player_cache_path = (
            CACHE_DIR
            / team.team_name
            / f"{self.position.value}_{self.player_name.replace(' ', '')}.html"
        )

        if use_cache and player_cache_path.exists():
            player_cache_data = load_file_cache(
                player_cache_path,
                f"Loading cached stats for {self.position.value} {self.player_name}...",
            )
            doc = BeautifulSoup(player_cache_data, "html.parser")
        else:
            response = handle_request(
                player_url,
                f"Fetching stats for {self.position.value} {self.player_name}...",
            )
            if save_results:
                cache_file(
                    player_cache_path,
                    response.text,
                    f"Caching stats for {self.position.value} {self.player_name}...",
                )
            doc = BeautifulSoup(response.text, "html.parser")

        rushing_table = doc.find("table", id="rushing_and_receiving")
        if rushing_table is None:
            print(
                "Table not found under rushing_and_receiving, searching for receiving_and_rushing"
            )
            rushing_table = doc.find("table", id="receiving_and_rushing")
            search_str = "receiving_and_rushing"
        else:
            search_str = "rushing_and_receiving"

        if rushing_table is None:
            print("Rushing/receiving table does not exist... assuming player is rookie")
            self.projections_exist = True
            return

        assert type(rushing_table) == Tag
        rushing_table_body = fetch_sub_tag(rushing_table, "tbody")

        for year in index:
            rushing_row = rushing_table_body.find("tr", id=f"{search_str}.{year}")
            if rushing_row is None:
                print(f"{self.player_name} does not have data for {year}")
                continue

            assert type(rushing_row) == Tag
            self.historical_data.loc[year, "team"] = fetch_data_stat(
                rushing_row, "Team Name", "team_name_abbr", stat_default_value="NA"
            )
            self.historical_data.loc[year, "games played"] = fetch_data_stat(
                rushing_row,
                "Games Played",
                "games",
                stat_default_value="NA",
                stat_dtype=int,
            )
            self.historical_data.loc[year, "games started"] = fetch_data_stat(
                rushing_row,
                "Games Started",
                "games_started",
                stat_default_value="NA",
                stat_dtype=int,
            )
            self.historical_data.loc[year, "target share"] = (
                fetch_data_stat(
                    rushing_row,
                    "targets",
                    "targets",
                    stat_default_value=0,
                    stat_dtype=float,
                )
                / self.team.historical_data.loc[year, "pass_plays"]
                * 100
            )
            catch_percentage = fetch_data_stat(
                rushing_row,
                "catch percentage",
                "catch_pct",
                stat_default_value=0,
                stat_dtype=float,
            )
            self.historical_data.loc[year, "catch %"] = catch_percentage

            self.historical_data.loc[year, "yards/catch"] = fetch_data_stat(
                rushing_row,
                "yards per catch",
                "rec_yds_per_rec",
                stat_default_value=0,
                stat_dtype=float,
            )

            rec_yds = fetch_data_stat(
                rushing_row,
                "receiving yards",
                "rec_yds",
                stat_default_value=0,
                stat_dtype=float,
            )
            assert type(rec_yds) == float
            # Need to do this to avoid a divide by 0 error
            if rec_yds <= 0:
                self.historical_data.loc[year, "TDs/rec_yard"] = 0
            else:
                rec_td = fetch_data_stat(
                    rushing_row,
                    "receiving touchdowns",
                    "rec_td",
                    stat_default_value=0,
                    stat_dtype=int,
                )
                assert type(rec_td) == int
                self.historical_data.loc[year, "TDs/rec_yard"] = rec_td / rec_yds

            rushing_attempts = fetch_data_stat(
                rushing_row,
                "rushing attempts",
                "rush_att",
                stat_default_value=0,
                stat_dtype=float,
            )
            assert type(rushing_attempts) == float
            self.historical_data.loc[year, "rush share"] = (
                rushing_attempts
                / self.team.historical_data.loc[year, "run_plays"]
                * 100
            )

            self.historical_data.loc[year, "ypc"] = fetch_data_stat(
                rushing_row,
                "yards per carry",
                "rush_yds_per_att",
                stat_default_value=0,
                stat_dtype=float,
            )

            rush_yds = fetch_data_stat(
                rushing_row,
                "rushing yards",
                "rush_yds",
                stat_default_value=0,
                stat_dtype=float,
            )
            assert type(rush_yds) == float
            # Need to do this to avoid a divide by 0 error
            if rush_yds <= 0:
                self.historical_data.loc[year, "tds/rush_yard"] = 0
            else:
                rush_td = fetch_data_stat(
                    rushing_row,
                    "rushing touchdowns",
                    "rush_td",
                    stat_default_value=0,
                    stat_dtype=int,
                )
                assert type(rush_td) == int
                self.historical_data.loc[year, "tds/rush_yard"] = rush_td / rush_yds

        self.projections_exist = True

    def project(self):
        """Prompts the user to enter projections for the current season"""
        print(self.historical_data)
        print(
            "Average games played: ",
            self.historical_data["games played"].dropna(how="all").mean(),
        )
        print(
            "Average target share: ",
            self.historical_data["target share"].dropna(how="all").mean(),
        )
        print(
            "Average catch %: ",
            self.historical_data["catch %"].dropna(how="all").mean(),
        )
        print(
            "Average yards/catch: ",
            self.historical_data["yards/catch"].dropna(how="all").mean(),
        )
        print(
            "Average TDs/rec_yard: ",
            self.historical_data["TDs/rec_yard"].dropna(how="all").mean(),
        )
        print(
            "Average rush share: ",
            self.historical_data["rush share"].dropna(how="all").mean(),
        )
        print("Average ypc: ", self.historical_data["ypc"].dropna(how="all").mean())
        print(
            "Average tds/rush_yard: ",
            self.historical_data["tds/rush_yard"].dropna(how="all").mean(),
        )
        print()
        self.games_played = prompt_for_stat("games played", int, self.current_year)
        self.target_share = prompt_for_stat("target share", float, self.current_year)
        self.catch_percentage = prompt_for_stat(
            "catch percentage", float, self.current_year
        )
        self.yards_per_catch = prompt_for_stat(
            "yards per catch", float, self.current_year
        )
        self.tds_per_rec_yd = prompt_for_stat(
            "receiving TDs per receiving yd", float, self.current_year
        )
        self.rush_share = prompt_for_stat("rushing share", float, self.current_year)
        self.yards_per_carry = prompt_for_stat(
            "yards per carry", float, self.current_year
        )
        self.tds_per_rush_yd = prompt_for_stat(
            "Rushing TDs per rush yd", float, self.current_year
        )

    def save_projections(self, filename: str):
        """Saves your player projection to a sheet"""
        file_path = Path(filename)
        if file_path.parent.exists:
            wb = load_workbook(filename)
            if self.team.team_name.capitalize() not in wb.sheetnames:
                wb.create_sheet(self.team.team_name.capitalize())
                wb.save(filename)
            existing_data = pd.read_excel(
                filename, sheet_name=self.team.team_name.capitalize()
            )
            if "Player Name" in existing_data:
                current_index = existing_data["Player Name"].size + 1
            else:
                current_index = 2
        else:
            file_path.parent.mkdir(parents=True)
            current_index = 2

        formatted_data = pd.DataFrame()
        formatted_data.loc[current_index, "Pos"] = self.position.value
        formatted_data.loc[current_index, "Player Name"] = self.player_name
        formatted_data.loc[current_index, "Games Started"] = self.games_played
        formatted_data.loc[current_index, "Target Share"] = self.target_share
        formatted_data.loc[current_index, "Catch Percentage"] = self.catch_percentage
        formatted_data.loc[current_index, "Yards/Catch"] = self.yards_per_catch
        formatted_data.loc[current_index, "TDs/Receiving Yard"] = self.tds_per_rec_yd
        formatted_data.loc[current_index, "Rush Share"] = self.rush_share
        formatted_data.loc[current_index, "Yards/Carry"] = self.yards_per_carry
        formatted_data.loc[current_index, "TDs/Rush Yard"] = self.tds_per_rush_yd
        print(formatted_data)
        if "existing_data" in locals():
            df_combined = pd.concat([existing_data, formatted_data])
            print(df_combined)
            with pd.ExcelWriter(
                filename, engine="openpyxl", mode="a", if_sheet_exists="replace"
            ) as writer:
                print("Merging data in excel...")
                df_combined.to_excel(
                    writer, sheet_name=self.team.team_name.capitalize(), index=False
                )
        else:
            with pd.ExcelWriter(
                filename, engine="openpyxl", mode="a", if_sheet_exists="replace"
            ) as writer:
                print("Saving new data in excel...")
                formatted_data.to_excel(
                    writer, sheet_name=self.team.team_name.capitalize(), index=False
                )
