from datetime import date
from pathlib import Path
from bs4 import BeautifulSoup, Tag
import pandas as pd
from openpyxl import load_workbook
from .teams import Team, Positions
from .caching import CACHE_DIR, load_file_cache, cache_file, handle_request
from .html_parsing import fetch_sub_tag, fetch_data_stat, fetch_specific_sub_tag
from .user_prompting import prompt_for_stat


class QB:
    """Projection process for quarterbacks"""

    # Position included to provide consistency across different implementations, not used for QB
    def __init__(
        self,
        team: Team,
        position: Positions,
        use_cache: bool = True,
        save_results: bool = True,
    ):
        """Searches through QBs on a team's current roster until you want to project one, then collects historical data for the past 3 years about that player"""
        self.team = team
        self.current_year = date.today().year
        index = pd.Index(range(self.current_year - 3, self.current_year + 1))
        columns = [
            "team",
            "games played",
            "games started",
            "pass_att",
            "int %",
            "pass td %",
            "comp %",
            "rush %",
            "ypc",
            "tds/rush_yard",
        ]
        self.historical_data = pd.DataFrame(index=index, columns=columns)

        player_name, player_url = team.get_player(position)

        if player_url is None or player_name is None:
            print(f"No projections will be done for this instance of {position.value}")
            self.projections_exist = False
            return

        self.player_name = player_name
        qb_path = CACHE_DIR / team.team_name / f"qb_{player_name.replace(' ', '')}.html"

        if use_cache and qb_path.exists():
            qb_data = load_file_cache(
                qb_path, f"Loading cached stats for QB {player_name}..."
            )
            doc = BeautifulSoup(qb_data, "html.parser")
        else:
            response = handle_request(
                player_url, f"Fetching stats for QB {player_name}..."
            )
            if save_results:
                cache_file(
                    qb_path, response.text, f"Caching stats for QB {player_name}"
                )
            doc = BeautifulSoup(response.text, "html.parser")

        # Need to handle the parsing for this one separately to handle rookies
        # Historical stat tables don't exist for rookies
        rushing_table = doc.find("table", id="rushing_and_receiving")
        if rushing_table is None:
            print("Rushing/receiving table does not exist... assuming player is rookie")
            self.projections_exist = True
            return
        assert type(rushing_table) == Tag

        passing_table = fetch_specific_sub_tag(doc, "table", "passing")
        passing_table_body = fetch_sub_tag(passing_table, "tbody")
        rushing_table_body = fetch_sub_tag(rushing_table, "tbody")

        # Slicing index here to avoid attempt to gather data for current year, which will always fail
        for year in index[0:3]:
            passing_row = passing_table_body.find("tr", id=f"passing.{year}")
            rushing_row = rushing_table_body.find(
                "tr", id=f"rushing_and_receiving.{year}"
            )

            if rushing_row is None or passing_row is None:
                print(f"{self.player_name} does not have data for {year}")
                continue

            assert type(passing_row) == Tag
            assert type(rushing_row) == Tag
            self.historical_data.loc[year, "team"] = fetch_data_stat(
                rushing_row, "Team Name", "team_name_abbr", stat_default_value="NA"
            )
            self.historical_data.loc[year, "games played"] = fetch_data_stat(
                rushing_row,
                "Games Played",
                "games",
                stat_default_value="NA",
                stat_dtype=int,
            )
            self.historical_data.loc[year, "games started"] = fetch_data_stat(
                rushing_row,
                "Games Started",
                "games_started",
                stat_default_value="NA",
                stat_dtype=int,
            )
            self.historical_data.loc[year, "pass_att"] = fetch_data_stat(
                passing_row, "Pass Attempts", "pass_att", stat_dtype=float
            )
            self.historical_data.loc[year, "int %"] = fetch_data_stat(
                passing_row, "Interception %", "pass_int_pct", stat_dtype=float
            )
            self.historical_data.loc[year, "pass td %"] = fetch_data_stat(
                passing_row, "Pass TD %", "pass_td_pct", stat_dtype=float
            )
            self.historical_data.loc[year, "comp %"] = fetch_data_stat(
                passing_row, "Completion %", "pass_cmp_pct", stat_dtype=float
            )
            self.historical_data.loc[year, "rush %"] = (
                fetch_data_stat(
                    rushing_row, "Rushing Attempts", "rush_att", stat_dtype=float
                )
                / self.team.historical_data.loc[year, "run_plays"]
                * 100
            )
            self.historical_data.loc[year, "ypc"] = fetch_data_stat(
                rushing_row, "Yards per carry", "rush_yds_per_att", stat_dtype=float
            )

            # Fetch rush TDs
            rush_tds = fetch_data_stat(
                rushing_row, "Rushing TDs", "rush_td", stat_dtype=int
            )
            assert type(rush_tds) == int

            # Fetch rush yds
            rush_yds = fetch_data_stat(
                rushing_row, "Rushing Yards", "rush_yds", stat_dtype=float
            )
            assert type(rush_yds) == float

            self.historical_data.loc[year, "tds/rush_yard"] = rush_tds / rush_yds
        self.projections_exist = True

    def project(self):
        """Prompts the user to enter projections for the current season"""
        print(self.historical_data)
        print(
            "Average games played: ",
            self.historical_data["games played"].dropna(how="all").mean(),
        )
        print(
            "Average int %: ",
            self.historical_data["int %"].drop(self.current_year).mean(),
        )
        print(
            "Average rush %: ",
            self.historical_data["rush %"].drop(self.current_year).mean(),
        )
        print(
            "Average ypc: ", self.historical_data["ypc"].drop(self.current_year).mean()
        )
        print(
            "Average tds/rush_yard: ",
            self.historical_data["tds/rush_yard"].drop(self.current_year).mean(),
        )
        print()

        self.games_played = prompt_for_stat("games played", int, self.current_year)
        self.int_percent = prompt_for_stat("int %", float, self.current_year)
        self.rush_percent = prompt_for_stat("rush %", float, self.current_year)
        self.ypc = prompt_for_stat("ypc", float, self.current_year)
        self.td_yard_ratio = prompt_for_stat(
            "Rushing TDs/yard", float, self.current_year
        )

    def save_projections(self, filename: str):
        """Saves your team-level projection to a sheet"""
        file_path = Path(filename)
        if file_path.parent.exists:
            wb = load_workbook(filename)
            if self.team.team_name.capitalize() not in wb.sheetnames:
                wb.create_sheet(self.team.team_name.capitalize())
                wb.save(filename)
            existing_data = pd.read_excel(
                filename, sheet_name=self.team.team_name.capitalize()
            )
            print(existing_data)
            if "Player Name" in existing_data:
                current_index = existing_data["Player Name"].size + 1
            else:
                current_index = 2
        else:
            file_path.parent.mkdir(parents=True)
            current_index = 2

        formatted_data = pd.DataFrame()
        formatted_data.loc[current_index, "Pos"] = "QB"
        formatted_data.loc[current_index, "Player Name"] = self.player_name
        formatted_data.loc[current_index, "Games Started"] = self.games_played
        formatted_data.loc[current_index, "Interception %"] = self.int_percent
        formatted_data.loc[current_index, "Rush Share"] = self.rush_percent
        formatted_data.loc[current_index, "Yards/Carry"] = self.ypc
        formatted_data.loc[current_index, "TDs/Rush Yard"] = self.td_yard_ratio
        print(formatted_data)
        if "existing_data" in locals():
            df_combined = pd.concat([existing_data, formatted_data])
            print(df_combined)
            with pd.ExcelWriter(
                filename, engine="openpyxl", mode="a", if_sheet_exists="replace"
            ) as writer:
                print("Merging data in excel...")
                df_combined.to_excel(
                    writer, sheet_name=self.team.team_name.capitalize(), index=False
                )
        else:
            with pd.ExcelWriter(
                filename, engine="openpyxl", mode="a", if_sheet_exists="replace"
            ) as writer:
                print("Saving new data in excel...")
                formatted_data.to_excel(
                    writer, sheet_name=self.team.team_name.capitalize(), index=False
                )
